from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def export_attendance_xlsx(
    out_path: Path, 
    discipline_name: str, 
    session_date: date, 
    students: list[dict]
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    ws["A1"] = "Дисциплина"
    ws["B1"] = discipline_name
    ws["A2"] = "Дата"
    ws["B2"] = session_date.isoformat()

    ws["A4"] = "№"
    ws["B4"] = "ФИО"
    ws["C4"] = "Присутствие"

    for i, st in enumerate(students, start=1):
        ws.cell(row=4 + i, column=1, value=i)
        ws.cell(row=4 + i, column=2, value=st["full_name"])
        ws.cell(row=4 + i, column=3, value="Да" if st["present"] else "Нет")

    for col in range(1, 4):
        max_len = 0
        for row in range(1, 5 + len(students)):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 50)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path

def export_discipline_xlsx(
    out_path: Path, 
    discipline_name: str,  
    discipline_attendance: list[dict],
    students: list[dict]
):
    wb = Workbook()
    
    ws = wb.active
    ws.title = "Посещение по дисциплине"

    ws["A1"] = "Дисциплина"
    ws["B1"] = discipline_name
    ws["B2"] = "Дата"
    ws["A3"] = "№"
    ws["B3"] = "ФИО"

    student_row_map = {}
    for i, st in enumerate(students, start=1):
        ws.cell(row=3 + i, column=1, value=i)
        ws.cell(row=3 + i, column=2, value=st["full_name"])
        student_row_map[st["full_name"]] = 3 + i

    for i, (date, attends) in enumerate(discipline_attendance.items(), start=1):
        col = 2 + i
        ws.cell(row=2, column=col, value=f"{date.day:02}.{date.month:02}")
        for student in attends:
            row = student_row_map.get(student)
            if row:
                ws.cell(row=row, column=col, value="✓")
    
    for col in range(1, 4):
        max_len = 0
        for row in range(1, 5 + len(students)):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 50)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


    